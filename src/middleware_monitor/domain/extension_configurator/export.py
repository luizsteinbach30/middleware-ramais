"""Exportação de ambientes do Configurador de Ramais em XLSX e PDF.

Gera um relatório por ambiente com o modelo do telefone, as configurações
(``config_padrao``) e a lista de ramais (IP, ramal, nome, etc.). Valores de
senha são **mascarados** (``****``) — o relatório é para conferência/entrega e
não deve vazar credenciais.

Libs: ``openpyxl`` (XLSX) e ``fpdf2`` (PDF) — ambas pure-Python, empacotáveis no
.exe sem dependências nativas.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO
from typing import Any

from sqlalchemy.orm import Session as DBSession

from . import repository as repo

# Campos de config_padrao exportados, na ordem, com rótulo amigável.
_CONFIG_LABELS: list[tuple[str, str]] = [
    ("sip_server", "Servidor SIP"),
    ("sip_transport", "Transporte SIP"),
    ("register_expiration", "Expiração de registro (s)"),
    ("ntp_server", "Servidor NTP"),
    ("timezone", "Fuso horário"),
    ("web_language", "Idioma web"),
    ("lcd_language", "Idioma do visor"),
    ("validar_conectividade", "Validar conectividade (ping)"),
    ("web_user", "Usuário web (atual)"),
    ("web_password", "Senha web (atual)"),
    ("nova_web_user", "Novo usuário web"),
    ("nova_web_password", "Nova senha web"),
    ("menu_password", "Senha do menu"),
    ("keylock_password", "Senha do bloqueio"),
    ("keylock_enable", "Bloqueio de teclado"),
    ("keylock_timeout", "Timeout do bloqueio (s)"),
]
_SECRET_HINTS = ("password", "senha", "pwd", "token")


def _mask(key: str, value: Any) -> str:
    if any(h in key.lower() for h in _SECRET_HINTS):
        return "****" if str(value or "") else "(vazio)"
    return str(value if value is not None else "")


def _format_function_keys(fks: list[dict[str, Any]]) -> str:
    parts: list[str] = []
    for fk in fks or []:
        key = fk.get("key", "?")
        tipo = fk.get("type", "?")
        if fk.get("value_source") == "linha":
            val = f"linha:{fk.get('value_field', 'numero_abreviado')}"
        else:
            val = str(fk.get("value_fixed", "") or "")
        label = fk.get("label", "")
        parts.append(f"{key}={tipo} {val} ({label})".strip())
    return "; ".join(parts)


@dataclass
class LineRow:
    ramal: str
    ip: str
    nome_visivel: str
    user_auth: str
    servidor_sip: str
    numero_abreviado: str


@dataclass
class EnvReport:
    id: str
    nome: str
    modelo_telefone: str
    config: list[tuple[str, str]] = field(default_factory=list)
    linhas: list[LineRow] = field(default_factory=list)


def build_reports(db: DBSession, env_ids: list[str]) -> list[EnvReport]:
    """Monta os relatórios dos ambientes pedidos (ignora ids inexistentes)."""
    reports: list[EnvReport] = []
    for env_id in env_ids:
        env = repo.get_environment(db, env_id)
        if env is None:
            continue
        cfg = repo.merged_config_padrao(env)
        config_pairs: list[tuple[str, str]] = [
            (label, _mask(key, cfg[key]))
            for key, label in _CONFIG_LABELS
            if key in cfg
        ]
        fks = _format_function_keys(cfg.get("function_keys") or [])
        if fks:
            config_pairs.append(("Teclas de função", fks))
        rows = [
            LineRow(
                ramal=ln.numero_ramal,
                ip=ln.ip,
                nome_visivel=ln.nome_visivel,
                user_auth=ln.user_auth,
                servidor_sip=ln.servidor_sip or str(cfg.get("sip_server", "")),
                numero_abreviado=ln.numero_abreviado,
            )
            for ln in repo.list_lines(db, env_id)
        ]
        reports.append(EnvReport(env.id, env.nome, env.modelo_telefone, config_pairs, rows))
    return reports


# --------------------------------------------------------------------- XLSX


def _sheet_title(nome: str, used: set[str]) -> str:
    """Título de aba válido: ≤31 chars, sem caracteres proibidos, único."""
    base = re.sub(r"[\[\]:*?/\\]", "-", nome).strip() or "Ambiente"
    base = base[:31]
    title = base
    i = 2
    while title.lower() in used:
        suffix = f" ({i})"
        title = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(title.lower())
    return title


def to_xlsx(reports: list[EnvReport]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    used: set[str] = set()
    for rep in reports:
        ws = wb.create_sheet(title=_sheet_title(rep.nome, used))
        ws["A1"] = f"Ambiente: {rep.nome}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Modelo do telefone: {rep.modelo_telefone}"
        r = 4
        ws.cell(r, 1, "Configurações").font = bold
        r += 1
        for label, value in rep.config:
            ws.cell(r, 1, label)
            ws.cell(r, 2, value)
            r += 1
        r += 1
        ws.cell(r, 1, f"Ramais ({len(rep.linhas)})").font = bold
        r += 1
        headers = ["Ramal", "IP", "Nome visível", "User auth", "Servidor SIP", "Nº abreviado"]
        for c, h in enumerate(headers, start=1):
            ws.cell(r, c, h).font = bold
        r += 1
        for ln in rep.linhas:
            for c, val in enumerate(
                [ln.ramal, ln.ip, ln.nome_visivel, ln.user_auth, ln.servidor_sip, ln.numero_abreviado],
                start=1,
            ):
                ws.cell(r, c, val)
            r += 1
        widths = [14, 16, 22, 16, 26, 14]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(1, c).column_letter].width = w
    if not reports:
        wb.create_sheet(title="Vazio")
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------- PDF


def _pdf_text(s: Any) -> str:
    """Fontes core do fpdf2 são latin-1; troca o que não couber por '?'."""
    return str(s if s is not None else "").encode("latin-1", "replace").decode("latin-1")


def to_pdf(reports: list[EnvReport]) -> bytes:
    from fpdf import FPDF

    from middleware_monitor import branding

    logo = branding.logo_for_pdf()
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    gerado = datetime.now().strftime("%d/%m/%Y %H:%M")
    for rep in reports:
        pdf.add_page()
        if logo is not None:
            try:
                pdf.image(str(logo), x=150, y=8, w=45)
            except Exception:  # imagem corrompida não deve quebrar o relatório
                pass
        pdf.set_font("Helvetica", "B", 15)
        pdf.cell(0, 9, _pdf_text(f"Ambiente: {rep.nome}"), new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 6, _pdf_text(f"Modelo do telefone: {rep.modelo_telefone}"), new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(120, 120, 120)
        pdf.cell(0, 5, _pdf_text(f"Gerado em {gerado}"), new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)
        pdf.ln(2)

        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 7, "Configuracoes", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 9)
        for label, value in rep.config:
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(55, 5, _pdf_text(label), new_x="RIGHT", new_y="TOP")
            pdf.set_font("Helvetica", "", 9)
            pdf.multi_cell(0, 5, _pdf_text(value), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 7, _pdf_text(f"Ramais ({len(rep.linhas)})"), new_x="LMARGIN", new_y="NEXT")
        widths = [22, 30, 38, 25, 45, 20]
        headers = ["Ramal", "IP", "Nome", "User auth", "Servidor SIP", "No. abrev"]
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(235, 235, 235)
        for w, h in zip(widths, headers, strict=True):
            pdf.cell(w, 6, _pdf_text(h), border=1, fill=True)
        pdf.ln(6)
        pdf.set_font("Helvetica", "", 8)
        for ln in rep.linhas:
            cells = [
                ln.ramal, ln.ip, ln.nome_visivel, ln.user_auth,
                ln.servidor_sip, ln.numero_abreviado,
            ]
            for w, val in zip(widths, cells, strict=True):
                txt = _pdf_text(val)
                # trunca para caber na coluna
                while txt and pdf.get_string_width(txt) > w - 2:
                    txt = txt[:-1]
                pdf.cell(w, 5, txt, border=1)
            pdf.ln(5)
        if not rep.linhas:
            pdf.set_font("Helvetica", "I", 8)
            pdf.cell(0, 5, "(sem ramais cadastrados)", new_x="LMARGIN", new_y="NEXT")
    if not reports:
        pdf.add_page()
        pdf.set_font("Helvetica", "", 11)
        pdf.cell(0, 8, "Nenhum ambiente para exportar.", new_x="LMARGIN", new_y="NEXT")
    out = pdf.output()
    return bytes(out)
